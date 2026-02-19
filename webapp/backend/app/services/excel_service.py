"""Excel 생성 서비스 - 밸브 리스트 + PIPE BOM"""
import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from copy import copy
from pathlib import Path
from collections import defaultdict, OrderedDict
import logging

logger = logging.getLogger(__name__)

# Styles
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DATA_FONT = Font(name="Arial", size=9)
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Piping class → material mapping
PIPING_CLASS_MAP = {
    "CS3": {"piping_spec": "ACS10B3", "body": "ASTM A536", "trim": "B62",
            "flange": "150 Lbs RF WN\n(A105)", "pipe_mat": "ASTM A53 Gr.B _ POLYETHYLENE LINING INSIDE"},
    "CS2": {"piping_spec": "ACS10B2", "body": "ASTM A536", "trim": "B62",
            "flange": "150 Lbs RF WN\n(A105)", "pipe_mat": "ASTM A53 Gr.B _ POLYETHYLENE LINING INSIDE"},
}

SSW_SPEC = {
    "piping_spec": "BCS21B3", "body": "ASTM A536", "trim": "B62",
    "flange": "150 Lbs RF WN\n(A105)", "pipe_mat": "ASTM A53 Gr.B _ POLYETHYLENE LINING INSIDE"
}

DESIGN_CONDITIONS = {
    "SW": {"press": 6.5, "temp": 60},
    "SSW": {"press": 10, "temp": 60},
    "CFW": {"press": 6.5, "temp": 90},
    "FW": {"press": 6.5, "temp": 60},
}


def _style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def _style_data(ws, row, max_col, font=None, fill=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font or DATA_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        if fill:
            cell.fill = fill


def _get_piping_spec(valve):
    fluid = valve.get("fluid", "")
    piping_class = valve.get("piping_class", "CS3")
    if fluid == "SW" and valve["tag"].startswith("SSW"):
        return SSW_SPEC["piping_spec"]
    return PIPING_CLASS_MAP.get(piping_class, PIPING_CLASS_MAP["CS3"])["piping_spec"]


def _get_material_info(valve):
    piping_class = valve.get("piping_class", "CS3")
    fluid = valve.get("fluid", "")
    if fluid == "SW" and valve["tag"].startswith("SSW"):
        return SSW_SPEC
    return PIPING_CLASS_MAP.get(piping_class, PIPING_CLASS_MAP["CS3"])


def _get_design_conditions(valve):
    tag = valve.get("tag", "")
    fluid = valve.get("fluid", "")
    if tag.startswith("SSW"):
        return DESIGN_CONDITIONS.get("SSW", {"press": 10, "temp": 60})
    return DESIGN_CONDITIONS.get(fluid, {"press": 6.5, "temp": 60})


def generate_valve_excel(valves: list[dict], output_path: str, template_path: str = None) -> str:
    """밸브 리스트 Excel 생성"""
    if template_path and Path(template_path).exists():
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active if "Manual" not in wb.sheetnames else wb["Manual"]

        # 참조 스타일 저장
        ref_styles = {}
        if ws.max_row >= 8:
            for cell in ws[8]:
                ref_styles[cell.column] = {
                    "font": copy(cell.font),
                    "alignment": copy(cell.alignment),
                    "border": copy(cell.border),
                    "fill": copy(cell.fill),
                }

        # 기존 데이터 클리어
        for row_num in range(8, ws.max_row + 1):
            for col in range(1, 27):
                ws.cell(row=row_num, column=col).value = None
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Valve List"
        ref_styles = {}

        # 헤더 생성
        headers = ["NO.", "TAG", "PIPING SPEC", "LOCATION", "DESCRIPTION", "FLUID",
                    "DESIGN PRESS", "DESIGN TEMP", "VALVE TYPE", "SUB TYPE",
                    "BODY", "TRIM", "SIZE", "PRESS RATING", "FLANGE",
                    "END IN", "END OUT", "PIPE MAT", "SCH IN", "SCH OUT",
                    "EXT BONNET", "CLASS CERT", "LS OPEN", "LS CLOSE", "LOCK DEV", "REMARK"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        _style_header(ws, 1, len(headers))

    # 밸브 분류
    manual_valves = sorted(
        [v for v in valves if v.get("valve_type") != "CONTROL"],
        key=lambda v: v["tag"]
    )
    control_valves = sorted(
        [v for v in valves if v.get("valve_type") == "CONTROL"],
        key=lambda v: v["tag"]
    )

    # Manual Valve 섹션
    start_row = 7 if template_path else 2
    if template_path:
        ws.cell(row=start_row, column=1, value="1. Manual Valve")

    for i, valve in enumerate(manual_valves):
        row = start_row + 1 + i
        mat_info = _get_material_info(valve)
        design = _get_design_conditions(valve)
        sch = valve.get("schedule", "STD")

        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=valve["tag"])
        ws.cell(row=row, column=3, value=_get_piping_spec(valve))
        ws.cell(row=row, column=4, value=valve.get("location", ""))
        ws.cell(row=row, column=5, value=valve.get("description", ""))
        ws.cell(row=row, column=6, value=valve.get("fluid", ""))
        ws.cell(row=row, column=7, value=design["press"])
        ws.cell(row=row, column=8, value=design["temp"])
        ws.cell(row=row, column=9, value=valve.get("valve_type", ""))
        ws.cell(row=row, column=10, value=valve.get("valve_subtype", ""))
        ws.cell(row=row, column=11, value=mat_info["body"])
        ws.cell(row=row, column=12, value=mat_info["trim"])
        ws.cell(row=row, column=13, value=int(valve["size"]) if valve.get("size", "").isdigit() else valve.get("size", ""))
        ws.cell(row=row, column=14, value="ANSI")
        ws.cell(row=row, column=15, value=mat_info["flange"])
        ws.cell(row=row, column=16, value="FLG")
        ws.cell(row=row, column=17, value="FLG")
        ws.cell(row=row, column=18, value=mat_info["pipe_mat"])
        ws.cell(row=row, column=19, value=sch)
        ws.cell(row=row, column=20, value=sch)
        ws.cell(row=row, column=21, value="-")
        ws.cell(row=row, column=22, value=3)
        ws.cell(row=row, column=23, value="-")
        ws.cell(row=row, column=24, value="-")
        ws.cell(row=row, column=25, value="-")
        ws.cell(row=row, column=26, value=f"Sheet {valve.get('sheet', '')}")

        for col in range(1, 27):
            cell = ws.cell(row=row, column=col)
            if col in ref_styles:
                cell.font = copy(ref_styles[col]["font"])
                cell.alignment = copy(ref_styles[col]["alignment"])
                cell.border = copy(ref_styles[col]["border"])

    # Control Valve 섹션
    ctrl_start = start_row + 1 + len(manual_valves) + 1
    ws.cell(row=ctrl_start, column=1, value="2. Control Valve")
    ws.cell(row=ctrl_start, column=1).font = Font(name="Arial", size=10, bold=True)

    for i, valve in enumerate(control_valves):
        row = ctrl_start + 1 + i
        mat_info = _get_material_info(valve)
        design = _get_design_conditions(valve)
        sch = valve.get("schedule", "STD")

        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=valve["tag"])
        ws.cell(row=row, column=3, value=_get_piping_spec(valve))
        ws.cell(row=row, column=4, value=valve.get("location", ""))
        ws.cell(row=row, column=5, value=valve.get("description", ""))
        ws.cell(row=row, column=6, value=valve.get("fluid", ""))
        ws.cell(row=row, column=7, value=design["press"])
        ws.cell(row=row, column=8, value=design["temp"])
        ws.cell(row=row, column=9, value="CONTROL")
        ws.cell(row=row, column=10, value=valve.get("valve_subtype", ""))
        ws.cell(row=row, column=11, value=mat_info["body"])
        ws.cell(row=row, column=12, value=mat_info["trim"])
        ws.cell(row=row, column=13, value=int(valve["size"]) if valve.get("size", "").isdigit() else valve.get("size", ""))
        ws.cell(row=row, column=14, value="ANSI")
        ws.cell(row=row, column=15, value=mat_info["flange"])
        ws.cell(row=row, column=16, value="FLG")
        ws.cell(row=row, column=17, value="FLG")
        ws.cell(row=row, column=18, value=mat_info["pipe_mat"])
        ws.cell(row=row, column=19, value=sch)
        ws.cell(row=row, column=20, value=sch)
        ws.cell(row=row, column=21, value="-")
        ws.cell(row=row, column=22, value=3)
        ws.cell(row=row, column=23, value="-")
        ws.cell(row=row, column=24, value="-")
        ws.cell(row=row, column=25, value="-")
        ws.cell(row=row, column=26, value=f"Sheet {valve.get('sheet', '')}")

        for col in range(1, 27):
            cell = ws.cell(row=row, column=col)
            if col in ref_styles:
                cell.font = copy(ref_styles[col]["font"])
                cell.alignment = copy(ref_styles[col]["alignment"])
                cell.border = copy(ref_styles[col]["border"])

    # 합계
    summary_row = ctrl_start + len(control_valves) + 2
    ws.cell(row=summary_row, column=1, value="TOTAL")
    ws.cell(row=summary_row, column=1).font = Font(name="Arial", size=10, bold=True)
    ws.cell(row=summary_row, column=2, value=f"Manual: {len(manual_valves)}, Control: {len(control_valves)}, Total: {len(valves)}")
    ws.cell(row=summary_row, column=2).font = Font(name="Arial", size=10, bold=True)

    wb.save(output_path)
    logger.info(f"Valve Excel saved: {output_path} ({len(valves)} valves)")
    return output_path


def generate_pipe_bom_excel(pages_data: list[dict], output_path: str) -> str:
    """PIPE BOM Excel 생성 (4개 시트)"""
    wb = openpyxl.Workbook()

    # === Sheet 1: Pipe Piece Summary ===
    ws1 = wb.active
    ws1.title = "Pipe Piece Summary"

    headers1 = ["NO.", "Page", "Pipe Piece No.", "Sub-pieces", "Weld Count",
                 "Loose Parts", "Pipe Lengths (mm)", "Total Length (mm)", "Revision Notes"]
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)
    _style_header(ws1, 1, len(headers1))

    row = 2
    total_welds = 0
    total_length = 0
    piece_no = 0

    for pd in pages_data:
        if not pd.get("pipe_pieces"):
            continue
        piece_no += 1
        dims = pd.get("dimensions_mm", [])
        total_dim = sum(dims) if dims else 0
        total_length += total_dim
        total_welds += pd.get("weld_count", 0)

        dims_str = ", ".join(str(d) for d in dims)
        other_dims = pd.get("other_dims", [])
        if other_dims:
            dims_str += (" + " if dims_str else "") + ", ".join(other_dims)

        rev = "; ".join(pd.get("revision_notes", []))

        ws1.cell(row=row, column=1, value=piece_no)
        ws1.cell(row=row, column=2, value=pd["page"])
        ws1.cell(row=row, column=3, value=", ".join(pd["pipe_pieces"]))
        ws1.cell(row=row, column=4, value=len(pd["pipe_pieces"]))
        ws1.cell(row=row, column=5, value=pd.get("weld_count", 0))
        ws1.cell(row=row, column=6, value="Yes" if pd.get("has_loose") else "-")
        ws1.cell(row=row, column=7, value=dims_str if dims_str else "-")
        ws1.cell(row=row, column=8, value=total_dim if total_dim > 0 else "-")
        ws1.cell(row=row, column=9, value=rev if rev else "-")
        _style_data(ws1, row, len(headers1))
        row += 1

    # 합계
    ws1.cell(row=row + 1, column=1, value="TOTAL")
    ws1.cell(row=row + 1, column=4, value=sum(len(p.get("pipe_pieces", [])) for p in pages_data))
    ws1.cell(row=row + 1, column=5, value=total_welds)
    ws1.cell(row=row + 1, column=8, value=total_length if total_length > 0 else "-")
    _style_data(ws1, row + 1, len(headers1),
                font=Font(name="Arial", size=10, bold=True),
                fill=PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"))

    for i, w in enumerate([6, 6, 45, 10, 10, 10, 30, 15, 40], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # === Sheet 2: Weld Item Detail ===
    ws2 = wb.create_sheet("Weld Item Detail")
    headers2 = ["NO.", "Page", "Pipe Piece", "Item No.", "Item Type", "Notes"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    _style_header(ws2, 1, len(headers2))

    row = 2
    item_no = 0
    for pd in pages_data:
        for weld in pd.get("weld_items", []):
            item_no += 1
            piece_str = ", ".join(pd.get("pipe_pieces", []))
            item_type = "Field Fit Weld (+100mm)" if weld.startswith("FFW") else "Shop Weld"
            ws2.cell(row=row, column=1, value=item_no)
            ws2.cell(row=row, column=2, value=pd["page"])
            ws2.cell(row=row, column=3, value=piece_str)
            ws2.cell(row=row, column=4, value=weld)
            ws2.cell(row=row, column=5, value=item_type)
            ws2.cell(row=row, column=6, value="-")
            _style_data(ws2, row, len(headers2))
            row += 1

    for i, w in enumerate([6, 6, 45, 10, 25, 20], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # === Sheet 3: Weld Quantity Summary ===
    ws3 = wb.create_sheet("Weld Quantity Summary")

    piece_summary = OrderedDict()
    for pd in pages_data:
        for pp in pd.get("pipe_pieces", []):
            base = pp.rsplit("-", 1)[0]
            if base not in piece_summary:
                piece_summary[base] = {
                    "sub_pieces": [], "shop_welds": 0, "field_welds": 0,
                    "total_welds": 0, "pages": set(), "dims": [], "has_loose": False,
                }
            piece_summary[base]["sub_pieces"].append(pp)
            piece_summary[base]["pages"].add(pd["page"])
            if pd.get("has_loose"):
                piece_summary[base]["has_loose"] = True

        if pd.get("pipe_pieces"):
            base = pd["pipe_pieces"][0].rsplit("-", 1)[0]
            for w in pd.get("weld_items", []):
                if w.startswith("FFW"):
                    piece_summary[base]["field_welds"] += 1
                else:
                    piece_summary[base]["shop_welds"] += 1
                piece_summary[base]["total_welds"] += 1
            for d in pd.get("dimensions_mm", []):
                piece_summary[base]["dims"].append(d)

    headers3 = ["NO.", "Pipe Piece Base", "Sub-piece Count", "Shop Welds",
                 "Field Welds", "Total Welds", "Has Loose", "Pipe Lengths (mm)",
                 "Total Length (mm)", "Pages"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    _style_header(ws3, 1, len(headers3))

    row = 2
    grand_shop = grand_field = grand_total = grand_length = 0

    for idx, (base, info) in enumerate(sorted(piece_summary.items()), 1):
        total_len = sum(info["dims"]) if info["dims"] else 0
        grand_shop += info["shop_welds"]
        grand_field += info["field_welds"]
        grand_total += info["total_welds"]
        grand_length += total_len

        ws3.cell(row=row, column=1, value=idx)
        ws3.cell(row=row, column=2, value=base)
        ws3.cell(row=row, column=3, value=len(set(info["sub_pieces"])))
        ws3.cell(row=row, column=4, value=info["shop_welds"])
        ws3.cell(row=row, column=5, value=info["field_welds"])
        ws3.cell(row=row, column=6, value=info["total_welds"])
        ws3.cell(row=row, column=7, value="Yes" if info["has_loose"] else "-")
        ws3.cell(row=row, column=8, value=", ".join(str(d) for d in info["dims"]) if info["dims"] else "-")
        ws3.cell(row=row, column=9, value=total_len if total_len > 0 else "-")
        ws3.cell(row=row, column=10, value=", ".join(str(p) for p in sorted(info["pages"])))
        _style_data(ws3, row, len(headers3))
        row += 1

    ws3.cell(row=row + 1, column=1, value="TOTAL")
    ws3.cell(row=row + 1, column=3, value=sum(len(set(v["sub_pieces"])) for v in piece_summary.values()))
    ws3.cell(row=row + 1, column=4, value=grand_shop)
    ws3.cell(row=row + 1, column=5, value=grand_field)
    ws3.cell(row=row + 1, column=6, value=grand_total)
    ws3.cell(row=row + 1, column=9, value=grand_length if grand_length > 0 else "-")
    _style_data(ws3, row + 1, len(headers3),
                font=Font(name="Arial", size=10, bold=True),
                fill=PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"))

    for i, w in enumerate([6, 18, 15, 12, 12, 12, 10, 30, 15, 15], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # === Sheet 4: Statistics ===
    ws4 = wb.create_sheet("Statistics")
    stats = [
        ("PIPE BOM STATISTICS", ""),
        ("", ""),
        ("Total Pages", len(pages_data)),
        ("Total Pipe Pieces", sum(len(p.get("pipe_pieces", [])) for p in pages_data)),
        ("Unique Base Pieces", len(piece_summary)),
        ("", ""),
        ("WELD SUMMARY", ""),
        ("Total Shop Welds", grand_shop),
        ("Total Field Welds (FFW)", grand_field),
        ("Total Welds", grand_total),
        ("", ""),
        ("PIPE LENGTH SUMMARY", ""),
        ("Total Measured Length (mm)", grand_length),
        ("Total Measured Length (m)", round(grand_length / 1000, 2) if grand_length else 0),
        ("", ""),
        ("LOOSE PARTS", ""),
        ("Pages with Loose Parts", sum(1 for p in pages_data if p.get("has_loose"))),
    ]

    for r, (label, value) in enumerate(stats, 1):
        ws4.cell(row=r, column=1, value=label)
        ws4.cell(row=r, column=2, value=value)
        if label and not value and label == label.upper():
            ws4.cell(row=r, column=1).font = Font(name="Arial", size=11, bold=True)
        else:
            ws4.cell(row=r, column=1).font = Font(name="Arial", size=10)
            ws4.cell(row=r, column=2).font = Font(name="Arial", size=10)

    ws4.column_dimensions["A"].width = 35
    ws4.column_dimensions["B"].width = 50

    wb.save(output_path)
    logger.info(f"Pipe BOM Excel saved: {output_path}")
    return output_path


# ─── VLM BOM 정밀 Excel 보고서 ─────────────────────────────

ACCENT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
WARN_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
MATCH_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # green
MISMATCH_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")  # red/pink
BOM_ONLY_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # yellow
DRAWING_ONLY_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")  # blue
NA_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # gray

STATUS_FILL_MAP = {
    "MATCH": MATCH_FILL,
    "MISMATCH": MISMATCH_FILL,
    "BOM_ONLY": BOM_ONLY_FILL,
    "DRAWING_ONLY": DRAWING_ONLY_FILL,
    "N/A": NA_FILL,
}


def generate_vlm_bom_excel(vlm_data: list[dict], output_path: str) -> str:
    """VLM 분석 기반 정밀 PIPE BOM Excel 생성 (7개 시트)"""
    wb = openpyxl.Workbook()

    # === Sheet 1: BOM Item List (전체 품목 + 비교 결과) ===
    ws1 = wb.active
    ws1.title = "BOM Item List"
    h1 = ["Page", "Drawing No.", "Line No.", "Code", "Qty", "Size",
          "Description", "Material Spec", "Weight (kg)", "Remarks",
          "Drawing Qty", "Match Status", "Diff"]
    for col, h in enumerate(h1, 1):
        ws1.cell(row=1, column=col, value=h)
    _style_header(ws1, 1, len(h1))

    row = 2
    total_items = 0
    total_weight = 0
    for page_data in vlm_data:
        page = page_data.get("page", 0)
        dwg_no = page_data.get("drawing_number", "")
        line_no = page_data.get("line_no", "") or (page_data.get("drawing_info", {}) or {}).get("line_no", "")

        # 비교 데이터 (letter_code 기준 lookup)
        comparison = page_data.get("comparison", {})
        comp_items_map = {}
        for ci in comparison.get("comparison_items", []):
            letter = ci.get("bom_letter", "")
            if letter:
                comp_items_map[letter] = ci

        for item in page_data.get("bom_table", []):
            total_items += 1
            code = item.get("letter_code", "") or item.get("item_no", "")
            wt = item.get("weight_kg", 0)
            if isinstance(wt, (int, float)) and wt > 0:
                total_weight += wt
            ws1.cell(row=row, column=1, value=page)
            ws1.cell(row=row, column=2, value=dwg_no)
            ws1.cell(row=row, column=3, value=line_no)
            ws1.cell(row=row, column=4, value=code)
            ws1.cell(row=row, column=5, value=item.get("quantity", ""))
            ws1.cell(row=row, column=6, value=item.get("size_inches", item.get("size", "")))
            ws1.cell(row=row, column=7, value=item.get("description", ""))
            ws1.cell(row=row, column=8, value=item.get("material_spec", item.get("material", "")))
            ws1.cell(row=row, column=9, value=wt if wt else "")
            ws1.cell(row=row, column=10, value=item.get("remarks", ""))

            # 비교 결과 컬럼
            ci = comp_items_map.get(code, {})
            match_status = ci.get("match_status", "")
            ws1.cell(row=row, column=11, value=ci.get("drawing_quantity", ""))
            ws1.cell(row=row, column=12, value=match_status)
            ws1.cell(row=row, column=13, value=ci.get("quantity_diff", "") if ci.get("quantity_diff") else "")

            _style_data(ws1, row, len(h1))

            # 비교 상태별 색상
            fill = STATUS_FILL_MAP.get(match_status)
            if fill:
                for c in range(11, 14):
                    ws1.cell(row=row, column=c).fill = fill

            row += 1

    for i, w in enumerate([6, 16, 8, 6, 8, 8, 40, 30, 10, 15, 10, 12, 6], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # === Sheet 2: Pipe Pieces ===
    ws2 = wb.create_sheet("Pipe Pieces")
    h2 = ["Page", "Drawing No.", "Pipe Group", "Piece ID", "Size", "Schedule", "Material"]
    for col, h in enumerate(h2, 1):
        ws2.cell(row=1, column=col, value=h)
    _style_header(ws2, 1, len(h2))

    row = 2
    for page_data in vlm_data:
        page = page_data.get("page", 0)
        dwg_no = page_data.get("drawing_number", "")
        pg = page_data.get("pipe_group", "")
        for pp in page_data.get("pipe_pieces", []):
            if isinstance(pp, dict):
                ws2.cell(row=row, column=1, value=page)
                ws2.cell(row=row, column=2, value=dwg_no)
                ws2.cell(row=row, column=3, value=pg)
                ws2.cell(row=row, column=4, value=pp.get("id", ""))
                ws2.cell(row=row, column=5, value=pp.get("size", ""))
                ws2.cell(row=row, column=6, value=pp.get("schedule", ""))
                ws2.cell(row=row, column=7, value=pp.get("material", ""))
            else:
                ws2.cell(row=row, column=1, value=page)
                ws2.cell(row=row, column=2, value=dwg_no)
                ws2.cell(row=row, column=3, value=pg)
                ws2.cell(row=row, column=4, value=str(pp))
            _style_data(ws2, row, len(h2))
            row += 1

    for i, w in enumerate([6, 15, 12, 15, 8, 10, 15], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # === Sheet 3: Components (Valves + Fittings) ===
    ws3 = wb.create_sheet("Components")
    h3 = ["Page", "Drawing No.", "Type", "Sub-type", "Size", "Tag", "Description", "Qty"]
    for col, h in enumerate(h3, 1):
        ws3.cell(row=1, column=col, value=h)
    _style_header(ws3, 1, len(h3))

    row = 2
    valve_count = 0
    fitting_count = 0
    for page_data in vlm_data:
        page = page_data.get("page", 0)
        dwg_no = page_data.get("drawing_number", "")
        for comp in page_data.get("components", []):
            ctype = comp.get("type", "")
            if ctype == "valve":
                valve_count += comp.get("quantity", 1)
            elif ctype == "fitting":
                fitting_count += comp.get("quantity", 1)
            ws3.cell(row=row, column=1, value=page)
            ws3.cell(row=row, column=2, value=dwg_no)
            ws3.cell(row=row, column=3, value=ctype.upper())
            ws3.cell(row=row, column=4, value=comp.get("subtype", ""))
            ws3.cell(row=row, column=5, value=comp.get("size", ""))
            ws3.cell(row=row, column=6, value=comp.get("tag", ""))
            ws3.cell(row=row, column=7, value=comp.get("description", ""))
            ws3.cell(row=row, column=8, value=comp.get("quantity", 1))
            fill = ACCENT_FILL if ctype == "valve" else None
            _style_data(ws3, row, len(h3), fill=fill)
            row += 1

    for i, w in enumerate([6, 15, 10, 18, 8, 12, 35, 5], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # === Sheet 4: Weld Points ===
    ws4 = wb.create_sheet("Weld Points")
    h4 = ["Page", "Drawing No.", "Weld ID", "Weld Type", "Notes"]
    for col, h in enumerate(h4, 1):
        ws4.cell(row=1, column=col, value=h)
    _style_header(ws4, 1, len(h4))

    row = 2
    shop_welds = 0
    field_welds = 0
    for page_data in vlm_data:
        page = page_data.get("page", 0)
        dwg_no = page_data.get("drawing_number", "")
        for wp in page_data.get("weld_points", []):
            wid = wp.get("id", "") if isinstance(wp, dict) else str(wp)
            wtype = wp.get("type", "shop_weld") if isinstance(wp, dict) else (
                "field_fit_weld" if "FFW" in str(wp).upper() else "shop_weld")
            if "field" in wtype.lower():
                field_welds += 1
            else:
                shop_welds += 1
            ws4.cell(row=row, column=1, value=page)
            ws4.cell(row=row, column=2, value=dwg_no)
            ws4.cell(row=row, column=3, value=wid)
            ws4.cell(row=row, column=4, value=wtype)
            ws4.cell(row=row, column=5, value="")
            fill = WARN_FILL if "field" in wtype.lower() else None
            _style_data(ws4, row, len(h4), fill=fill)
            row += 1

    for i, w in enumerate([6, 15, 10, 20, 20], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    # === Sheet 5: Dimensions ===
    ws5 = wb.create_sheet("Dimensions")
    h5 = ["Page", "From Point", "To Point", "Length (mm)", "Direction"]
    for col, h in enumerate(h5, 1):
        ws5.cell(row=1, column=col, value=h)
    _style_header(ws5, 1, len(h5))

    row = 2
    total_length = 0
    for page_data in vlm_data:
        page = page_data.get("page", 0)
        for dim in page_data.get("dimensions_mm", []):
            if isinstance(dim, dict):
                length = dim.get("length_mm", 0)
                total_length += length if isinstance(length, (int, float)) else 0
                ws5.cell(row=row, column=1, value=page)
                ws5.cell(row=row, column=2, value=dim.get("from_point", ""))
                ws5.cell(row=row, column=3, value=dim.get("to_point", ""))
                ws5.cell(row=row, column=4, value=length)
                ws5.cell(row=row, column=5, value=dim.get("direction", ""))
            else:
                total_length += dim if isinstance(dim, (int, float)) else 0
                ws5.cell(row=row, column=1, value=page)
                ws5.cell(row=row, column=4, value=dim)
            _style_data(ws5, row, len(h5))
            row += 1

    for i, w in enumerate([6, 12, 12, 12, 12], 1):
        ws5.column_dimensions[get_column_letter(i)].width = w

    # === Sheet 6: Cut Lengths ===
    ws6 = wb.create_sheet("Cut Lengths")
    h6 = ["Page", "Drawing No.", "Line No.", "Cut No.", "Length (mm)"]
    for col, h in enumerate(h6, 1):
        ws6.cell(row=1, column=col, value=h)
    _style_header(ws6, 1, len(h6))

    row = 2
    total_cut_length = 0
    for page_data in vlm_data:
        page = page_data.get("page", 0)
        dwg_no = page_data.get("drawing_number", "")
        line_no = page_data.get("line_no", "") or (page_data.get("drawing_info", {}) or {}).get("line_no", "")
        for cut in page_data.get("cut_lengths", []):
            if isinstance(cut, dict):
                length = cut.get("length_mm", 0)
                total_cut_length += length if isinstance(length, (int, float)) else 0
                ws6.cell(row=row, column=1, value=page)
                ws6.cell(row=row, column=2, value=dwg_no)
                ws6.cell(row=row, column=3, value=line_no)
                ws6.cell(row=row, column=4, value=cut.get("cut_no", ""))
                ws6.cell(row=row, column=5, value=length)
                _style_data(ws6, row, len(h6))
                row += 1

    for i, w in enumerate([6, 16, 8, 8, 12], 1):
        ws6.column_dimensions[get_column_letter(i)].width = w

    # === Sheet 7: Drawing Index ===
    ws7_idx = wb.create_sheet("Drawing Index")
    h7 = ["Page", "Drawing No.", "Line No.", "Pipe No.", "Line Description",
          "Pipe Pieces", "Shop Welds", "Field Welds", "BOM Items", "Cut Lengths",
          "Total Weight (kg)", "Revision"]
    for col, h in enumerate(h7, 1):
        ws7_idx.cell(row=1, column=col, value=h)
    _style_header(ws7_idx, 1, len(h7))

    row = 2
    for page_data in vlm_data:
        di = page_data.get("drawing_info", {}) or {}
        ws7_idx.cell(row=row, column=1, value=page_data.get("page", 0))
        ws7_idx.cell(row=row, column=2, value=page_data.get("drawing_number", ""))
        ws7_idx.cell(row=row, column=3, value=page_data.get("line_no", "") or di.get("line_no", ""))
        ws7_idx.cell(row=row, column=4, value=page_data.get("pipe_no", "") or di.get("pipe_no", ""))
        ws7_idx.cell(row=row, column=5, value=page_data.get("line_description", "") or di.get("line_description", ""))
        ws7_idx.cell(row=row, column=6, value=len(page_data.get("pipe_pieces", [])))
        sw = sum(1 for w in page_data.get("weld_points", [])
                 if isinstance(w, dict) and "field" not in w.get("type", "").lower())
        fw = sum(1 for w in page_data.get("weld_points", [])
                 if isinstance(w, dict) and "field" in w.get("type", "").lower())
        ws7_idx.cell(row=row, column=7, value=sw)
        ws7_idx.cell(row=row, column=8, value=fw)
        ws7_idx.cell(row=row, column=9, value=len(page_data.get("bom_table", [])))
        ws7_idx.cell(row=row, column=10, value=len(page_data.get("cut_lengths", [])))
        bom_wt = sum(item.get("weight_kg", 0) for item in page_data.get("bom_table", [])
                     if isinstance(item.get("weight_kg"), (int, float)))
        ws7_idx.cell(row=row, column=11, value=bom_wt if bom_wt else "")
        ws7_idx.cell(row=row, column=12, value=di.get("revision", ""))
        _style_data(ws7_idx, row, len(h7))
        row += 1

    for i, w in enumerate([6, 16, 8, 12, 35, 10, 10, 10, 10, 10, 12, 8], 1):
        ws7_idx.column_dimensions[get_column_letter(i)].width = w

    # === Sheet 8: Summary Statistics ===
    ws_summary = wb.create_sheet("Summary")
    stats = [
        ("VLM PIPE BOM EXTRACTION REPORT", ""),
        ("", ""),
        ("OVERVIEW", ""),
        ("Total Pages Analyzed", len(vlm_data)),
        ("Pages with Data", sum(1 for r in vlm_data if r.get("pipe_pieces") or r.get("bom_table"))),
        ("Drawing Analysis Success", sum(1 for r in vlm_data if r.get("drawing_analysis_ok"))),
        ("Table Analysis Success", sum(1 for r in vlm_data if r.get("table_analysis_ok"))),
        ("", ""),
        ("PIPE PIECES", ""),
        ("Total Pipe Pieces", sum(len(r.get("pipe_pieces", [])) for r in vlm_data)),
        ("", ""),
        ("COMPONENTS", ""),
        ("Total Valves", valve_count),
        ("Total Fittings", fitting_count),
        ("Total Other Components", sum(
            sum(1 for c in r.get("components", []) if c.get("type") not in ("valve", "fitting"))
            for r in vlm_data)),
        ("", ""),
        ("WELDING", ""),
        ("Total Shop Welds", shop_welds),
        ("Total Field Fit Welds", field_welds),
        ("Total Welds", shop_welds + field_welds),
        ("", ""),
        ("DIMENSIONS", ""),
        ("Total Dimension Entries", sum(len(r.get("dimensions_mm", [])) for r in vlm_data)),
        ("Total Pipe Length (mm)", total_length),
        ("Total Pipe Length (m)", round(total_length / 1000, 2) if total_length else 0),
        ("", ""),
        ("CUT LENGTHS", ""),
        ("Total Cut Entries", sum(len(r.get("cut_lengths", [])) for r in vlm_data)),
        ("Total Cut Length (mm)", total_cut_length),
        ("Total Cut Length (m)", round(total_cut_length / 1000, 2) if total_cut_length else 0),
        ("", ""),
        ("BOM TABLE", ""),
        ("Total BOM Items", total_items),
        ("Total Weight (kg)", round(total_weight, 1)),
    ]

    for r, (label, value) in enumerate(stats, 1):
        ws_summary.cell(row=r, column=1, value=label)
        ws_summary.cell(row=r, column=2, value=value)
        if label and not value and label == label.upper():
            ws_summary.cell(row=r, column=1).font = Font(name="Arial", size=11, bold=True)
        else:
            ws_summary.cell(row=r, column=1).font = Font(name="Arial", size=10)
            ws_summary.cell(row=r, column=2).font = Font(name="Arial", size=10)

    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 50

    # === Sheet 9: BOM Comparison (비교 상세) ===
    has_comparison = any(pd.get("comparison") for pd in vlm_data)
    if has_comparison:
        ws_comp = wb.create_sheet("BOM Comparison")
        hc = ["Page", "Drawing No.", "BOM Code", "BOM Description", "BOM Qty",
              "BOM Size", "Drawing Component", "Drawing Qty", "Status", "Diff", "Notes"]
        for col, h in enumerate(hc, 1):
            ws_comp.cell(row=1, column=col, value=h)
        _style_header(ws_comp, 1, len(hc))

        row = 2
        for page_data in vlm_data:
            comparison = page_data.get("comparison", {})
            if not comparison:
                continue
            page = comparison.get("page", page_data.get("page", 0))
            dwg_no = comparison.get("drawing_number", page_data.get("drawing_number", ""))
            for ci in comparison.get("comparison_items", []):
                ws_comp.cell(row=row, column=1, value=page)
                ws_comp.cell(row=row, column=2, value=dwg_no)
                ws_comp.cell(row=row, column=3, value=ci.get("bom_letter", ""))
                ws_comp.cell(row=row, column=4, value=ci.get("bom_description", ""))
                ws_comp.cell(row=row, column=5, value=ci.get("bom_quantity", ""))
                ws_comp.cell(row=row, column=6, value=ci.get("bom_size", ""))
                ws_comp.cell(row=row, column=7, value=ci.get("drawing_component", ""))
                ws_comp.cell(row=row, column=8, value=ci.get("drawing_quantity", ""))
                status = ci.get("match_status", "")
                ws_comp.cell(row=row, column=9, value=status)
                ws_comp.cell(row=row, column=10, value=ci.get("quantity_diff", ""))
                ws_comp.cell(row=row, column=11, value=ci.get("notes", ""))
                _style_data(ws_comp, row, len(hc))

                fill = STATUS_FILL_MAP.get(status)
                if fill:
                    for c in range(1, len(hc) + 1):
                        ws_comp.cell(row=row, column=c).fill = fill
                row += 1

        for i, w in enumerate([6, 16, 8, 35, 8, 8, 20, 10, 12, 6, 30], 1):
            ws_comp.column_dimensions[get_column_letter(i)].width = w

        # === Sheet 10: Comparison Summary (비교 요약) ===
        ws_cs = wb.create_sheet("Comparison Summary")
        hcs = ["Page", "Drawing No.", "Line No.", "BOM Items", "Comparable",
               "Matched", "Mismatched", "BOM Only", "Drawing Only", "Match Rate (%)"]
        for col, h in enumerate(hcs, 1):
            ws_cs.cell(row=1, column=col, value=h)
        _style_header(ws_cs, 1, len(hcs))

        row = 2
        tot_matched = tot_mismatched = tot_bom_only = tot_drawing_only = tot_comparable = 0
        for page_data in vlm_data:
            comparison = page_data.get("comparison", {})
            if not comparison:
                continue
            summary = comparison.get("summary", {})
            matched = summary.get("matched", 0)
            mismatched = summary.get("mismatched", 0)
            bom_only = summary.get("bom_only", 0)
            drawing_only = summary.get("drawing_only", 0)
            comparable = summary.get("comparable_items", 0)
            tot_matched += matched
            tot_mismatched += mismatched
            tot_bom_only += bom_only
            tot_drawing_only += drawing_only
            tot_comparable += comparable

            ws_cs.cell(row=row, column=1, value=comparison.get("page", 0))
            ws_cs.cell(row=row, column=2, value=comparison.get("drawing_number", ""))
            ws_cs.cell(row=row, column=3, value=comparison.get("line_no", ""))
            ws_cs.cell(row=row, column=4, value=summary.get("total_bom_items", 0))
            ws_cs.cell(row=row, column=5, value=comparable)
            ws_cs.cell(row=row, column=6, value=matched)
            ws_cs.cell(row=row, column=7, value=mismatched)
            ws_cs.cell(row=row, column=8, value=bom_only)
            ws_cs.cell(row=row, column=9, value=drawing_only)
            ws_cs.cell(row=row, column=10, value=summary.get("match_rate", 0))
            _style_data(ws_cs, row, len(hcs))

            # 낮은 일치율 강조
            if summary.get("match_rate", 100) < 50:
                ws_cs.cell(row=row, column=10).fill = MISMATCH_FILL
            row += 1

        # 합계 행
        ws_cs.cell(row=row + 1, column=1, value="TOTAL")
        ws_cs.cell(row=row + 1, column=5, value=tot_comparable)
        ws_cs.cell(row=row + 1, column=6, value=tot_matched)
        ws_cs.cell(row=row + 1, column=7, value=tot_mismatched)
        ws_cs.cell(row=row + 1, column=8, value=tot_bom_only)
        ws_cs.cell(row=row + 1, column=9, value=tot_drawing_only)
        overall_rate = round(tot_matched / max(1, tot_comparable) * 100, 1)
        ws_cs.cell(row=row + 1, column=10, value=overall_rate)
        _style_data(ws_cs, row + 1, len(hcs),
                    font=Font(name="Arial", size=10, bold=True),
                    fill=PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"))

        for i, w in enumerate([6, 16, 8, 10, 10, 10, 10, 10, 10, 12], 1):
            ws_cs.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    logger.info(f"VLM BOM Excel saved: {output_path} ({total_items} BOM items, "
                f"{valve_count} valves, {fitting_count} fittings)")
    return output_path


# ─── P&ID VLM 분석 Excel ─────────────────────────────

def generate_pid_analysis_excel(valves: list[dict], line_specs: list[dict],
                                 symbols_found: list[dict], output_path: str) -> str:
    """P&ID VLM 분석 결과 Excel 생성 (3개 시트)"""
    wb = openpyxl.Workbook()

    # === Sheet 1: Pipe & Valve List ===
    ws1 = wb.active
    ws1.title = "Pipe & Valve List"
    h1 = ["No", "Tag", "Symbol Type", "Valve Type", "Actuator", "Size",
          "Line Spec", "Piping Class", "Schedule", "Pressure Rating",
          "Material", "Fluid", "Description", "Sheet", "Source"]
    for col, h in enumerate(h1, 1):
        ws1.cell(row=1, column=col, value=h)
    _style_header(ws1, 1, len(h1))

    row = 2
    for i, v in enumerate(valves, 1):
        ws1.cell(row=row, column=1, value=i)
        ws1.cell(row=row, column=2, value=v.get("tag", ""))
        ws1.cell(row=row, column=3, value=v.get("valve_subtype", v.get("valve_type", "")))
        ws1.cell(row=row, column=4, value=v.get("valve_type", ""))
        ws1.cell(row=row, column=5, value=v.get("actuator", ""))
        ws1.cell(row=row, column=6, value=v.get("size", ""))
        ws1.cell(row=row, column=7, value=v.get("line_spec", ""))
        ws1.cell(row=row, column=8, value=v.get("piping_class", ""))
        ws1.cell(row=row, column=9, value=v.get("schedule", ""))
        ws1.cell(row=row, column=10, value=v.get("pressure_rating", ""))
        ws1.cell(row=row, column=11, value=v.get("material_code", ""))
        ws1.cell(row=row, column=12, value=v.get("fluid", ""))
        ws1.cell(row=row, column=13, value=v.get("description", ""))
        ws1.cell(row=row, column=14, value=v.get("sheet", ""))
        ws1.cell(row=row, column=15, value=v.get("source", ""))
        _style_data(ws1, row, len(h1))

        # VLM/Both 소스 강조
        source = v.get("source", "")
        if source == "vlm":
            ws1.cell(row=row, column=15).fill = ACCENT_FILL
        elif source == "both":
            ws1.cell(row=row, column=15).fill = PatternFill(
                start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        row += 1

    for i, w in enumerate([5, 12, 20, 12, 10, 6, 35, 10, 10, 12, 10, 6, 30, 6, 8], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # === Sheet 2: Line Specifications ===
    ws2 = wb.create_sheet("Line Specifications")
    h2 = ["No", "Full Spec", "Size", "System Code", "Line Number", "Tag",
          "Piping Class", "Schedule", "Pressure Rating", "Material", "Fluid", "Sheet"]
    for col, h in enumerate(h2, 1):
        ws2.cell(row=1, column=col, value=h)
    _style_header(ws2, 1, len(h2))

    row = 2
    for i, ls in enumerate(line_specs, 1):
        ws2.cell(row=row, column=1, value=i)
        ws2.cell(row=row, column=2, value=ls.get("full_spec", ""))
        ws2.cell(row=row, column=3, value=ls.get("size", ""))
        ws2.cell(row=row, column=4, value=ls.get("system_code", ""))
        ws2.cell(row=row, column=5, value=ls.get("line_number", ""))
        ws2.cell(row=row, column=6, value=ls.get("tag", ""))
        ws2.cell(row=row, column=7, value=ls.get("piping_class", ""))
        ws2.cell(row=row, column=8, value=ls.get("schedule", ""))
        ws2.cell(row=row, column=9, value=ls.get("pressure_rating", ""))
        ws2.cell(row=row, column=10, value=ls.get("material_code", ""))
        ws2.cell(row=row, column=11, value=ls.get("fluid", ""))
        ws2.cell(row=row, column=12, value=ls.get("sheet", ""))
        _style_data(ws2, row, len(h2))
        row += 1

    for i, w in enumerate([5, 35, 6, 12, 12, 12, 10, 10, 12, 10, 6, 6], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # === Sheet 3: Summary ===
    ws3 = wb.create_sheet("Summary")
    valve_by_type = defaultdict(int)
    for v in valves:
        vt = v.get("valve_type", "UNKNOWN")
        valve_by_type[vt] += 1

    system_count = defaultdict(int)
    for ls in line_specs:
        sc = ls.get("system_code", "UNKNOWN")
        system_count[sc] += 1

    stats = [
        ("P&ID ANALYSIS REPORT", ""),
        ("", ""),
        ("OVERVIEW", ""),
        ("Total Line Specifications", len(line_specs)),
        ("Total Valves", len(valves)),
        ("Total Symbols Found", len(symbols_found)),
        ("", ""),
        ("VALVES BY TYPE", ""),
    ]
    for vt, cnt in sorted(valve_by_type.items()):
        stats.append((f"  {vt}", cnt))
    stats.extend([
        ("", ""),
        ("LINE SPECS BY SYSTEM", ""),
    ])
    for sc, cnt in sorted(system_count.items()):
        stats.append((f"  {sc}", cnt))

    for r, (label, value) in enumerate(stats, 1):
        ws3.cell(row=r, column=1, value=label)
        ws3.cell(row=r, column=2, value=value)
        if label and not value and label == label.upper():
            ws3.cell(row=r, column=1).font = Font(name="Arial", size=11, bold=True)
        else:
            ws3.cell(row=r, column=1).font = Font(name="Arial", size=10)
            ws3.cell(row=r, column=2).font = Font(name="Arial", size=10)

    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 50

    wb.save(output_path)
    logger.info(f"P&ID analysis Excel saved: {output_path} "
                f"({len(valves)} valves, {len(line_specs)} line specs)")
    return output_path
