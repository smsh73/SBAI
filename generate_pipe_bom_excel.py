#!/usr/bin/env python3
"""PIPE BOM Excel 생성 - 46페이지 PDF에서 추출한 데이터를 엑셀로 출력"""
import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path

SBAI_DIR = Path(__file__).parent
BOM_DATA_PATH = SBAI_DIR / "pipe_bom_data.json"
OUTPUT_PATH = SBAI_DIR / "3. 260211-PIPE_BOM-OUTPUT.xlsx"

# Styles
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DATA_FONT = Font(name="Arial", size=9)
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def style_data(ws, row, max_col, font=None, fill=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font or DATA_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        if fill:
            cell.fill = fill


def main():
    with open(BOM_DATA_PATH) as f:
        pages_data = json.load(f)

    wb = openpyxl.Workbook()

    # ========================================
    # Sheet 1: Pipe Piece Summary (파이프 피스 요약)
    # ========================================
    ws1 = wb.active
    ws1.title = "Pipe Piece Summary"

    headers1 = ["NO.", "Page", "Pipe Piece No.", "Sub-pieces", "Weld Count",
                 "Loose Parts", "Pipe Lengths (mm)", "Total Length (mm)",
                 "Revision Notes"]
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)
    style_header(ws1, 1, len(headers1))

    row = 2
    total_welds = 0
    total_length = 0
    piece_no = 0

    for pd in pages_data:
        if not pd["pipe_pieces"]:
            continue
        piece_no += 1

        # Group pipe pieces by base number
        base_piece = pd["pipe_pieces"][0].rsplit("-", 1)[0] if pd["pipe_pieces"] else ""
        sub_count = len(pd["pipe_pieces"])
        dims = pd.get("dimensions_mm", [])
        other_dims = pd.get("other_dims", [])
        total_dim = sum(dims) if dims else 0
        total_length += total_dim
        total_welds += pd["weld_count"]

        dims_str = ", ".join(str(d) for d in dims)
        if other_dims:
            dims_str += (" + " if dims_str else "") + ", ".join(other_dims)

        rev = "; ".join(pd.get("revision_notes", []))

        ws1.cell(row=row, column=1, value=piece_no)
        ws1.cell(row=row, column=2, value=pd["page"])
        ws1.cell(row=row, column=3, value=", ".join(pd["pipe_pieces"]))
        ws1.cell(row=row, column=4, value=sub_count)
        ws1.cell(row=row, column=5, value=pd["weld_count"])
        ws1.cell(row=row, column=6, value="Yes" if pd.get("has_loose") else "-")
        ws1.cell(row=row, column=7, value=dims_str if dims_str else "-")
        ws1.cell(row=row, column=8, value=total_dim if total_dim > 0 else "-")
        ws1.cell(row=row, column=9, value=rev if rev else "-")
        style_data(ws1, row, len(headers1))
        row += 1

    # Summary row
    ws1.cell(row=row + 1, column=1, value="TOTAL")
    ws1.cell(row=row + 1, column=4, value=sum(len(p["pipe_pieces"]) for p in pages_data))
    ws1.cell(row=row + 1, column=5, value=total_welds)
    ws1.cell(row=row + 1, column=8, value=total_length if total_length > 0 else "-")
    style_data(ws1, row + 1, len(headers1),
               font=Font(name="Arial", size=10, bold=True),
               fill=PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"))

    # Column widths
    col_widths1 = [6, 6, 45, 10, 10, 10, 30, 15, 40]
    for i, w in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ========================================
    # Sheet 2: Weld Item Detail (용접 항목 상세)
    # ========================================
    ws2 = wb.create_sheet("Weld Item Detail")

    headers2 = ["NO.", "Page", "Pipe Piece", "Item No.", "Item Type", "Notes"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header(ws2, 1, len(headers2))

    row = 2
    item_no = 0
    for pd in pages_data:
        for weld in pd.get("weld_items", []):
            item_no += 1
            piece_str = ", ".join(pd["pipe_pieces"])

            # Determine item type
            if weld.startswith("FFW"):
                item_type = "Field Fit Weld (+100mm)"
            elif weld.startswith("W"):
                item_type = "Shop Weld"
            else:
                item_type = "Other"

            ws2.cell(row=row, column=1, value=item_no)
            ws2.cell(row=row, column=2, value=pd["page"])
            ws2.cell(row=row, column=3, value=piece_str)
            ws2.cell(row=row, column=4, value=weld)
            ws2.cell(row=row, column=5, value=item_type)
            ws2.cell(row=row, column=6, value="-")
            style_data(ws2, row, len(headers2))
            row += 1

    col_widths2 = [6, 6, 45, 10, 25, 20]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ========================================
    # Sheet 3: Weld Quantity Summary (용접 물량 취합)
    # ========================================
    ws3 = wb.create_sheet("Weld Quantity Summary")

    # Aggregate by pipe piece base number
    from collections import defaultdict, OrderedDict

    piece_summary = OrderedDict()
    for pd in pages_data:
        for pp in pd["pipe_pieces"]:
            base = pp.rsplit("-", 1)[0]
            if base not in piece_summary:
                piece_summary[base] = {
                    "sub_pieces": [],
                    "shop_welds": 0,
                    "field_welds": 0,
                    "total_welds": 0,
                    "pages": set(),
                    "dims": [],
                    "has_loose": False,
                }
            piece_summary[base]["sub_pieces"].append(pp)
            piece_summary[base]["pages"].add(pd["page"])
            if pd.get("has_loose"):
                piece_summary[base]["has_loose"] = True

        # Count weld types
        if pd["pipe_pieces"]:
            base = pd["pipe_pieces"][0].rsplit("-", 1)[0]
            for w in pd.get("weld_items", []):
                if w.startswith("FFW"):
                    piece_summary[base]["field_welds"] += 1
                else:
                    piece_summary[base]["shop_welds"] += 1
                piece_summary[base]["total_welds"] += 1

            for d in pd.get("dimensions_mm", []):
                piece_summary[base]["dims"].append(d)

    headers3 = ["NO.", "Pipe Piece Base", "Sub-piece Count", "Shop Welds",
                 "Field Welds", "Total Welds", "Has Loose", "Pipe Lengths (mm)",
                 "Total Length (mm)", "Pages"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    style_header(ws3, 1, len(headers3))

    row = 2
    grand_shop = 0
    grand_field = 0
    grand_total = 0
    grand_length = 0

    for idx, (base, info) in enumerate(sorted(piece_summary.items()), 1):
        total_len = sum(info["dims"]) if info["dims"] else 0
        grand_shop += info["shop_welds"]
        grand_field += info["field_welds"]
        grand_total += info["total_welds"]
        grand_length += total_len

        ws3.cell(row=row, column=1, value=idx)
        ws3.cell(row=row, column=2, value=base)
        ws3.cell(row=row, column=3, value=len(set(info["sub_pieces"])))
        ws3.cell(row=row, column=4, value=info["shop_welds"])
        ws3.cell(row=row, column=5, value=info["field_welds"])
        ws3.cell(row=row, column=6, value=info["total_welds"])
        ws3.cell(row=row, column=7, value="Yes" if info["has_loose"] else "-")
        ws3.cell(row=row, column=8, value=", ".join(str(d) for d in info["dims"]) if info["dims"] else "-")
        ws3.cell(row=row, column=9, value=total_len if total_len > 0 else "-")
        ws3.cell(row=row, column=10, value=", ".join(str(p) for p in sorted(info["pages"])))
        style_data(ws3, row, len(headers3))
        row += 1

    # Grand total
    ws3.cell(row=row + 1, column=1, value="TOTAL")
    ws3.cell(row=row + 1, column=3, value=sum(len(set(v["sub_pieces"])) for v in piece_summary.values()))
    ws3.cell(row=row + 1, column=4, value=grand_shop)
    ws3.cell(row=row + 1, column=5, value=grand_field)
    ws3.cell(row=row + 1, column=6, value=grand_total)
    ws3.cell(row=row + 1, column=9, value=grand_length if grand_length > 0 else "-")
    style_data(ws3, row + 1, len(headers3),
               font=Font(name="Arial", size=10, bold=True),
               fill=PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"))

    col_widths3 = [6, 18, 15, 12, 12, 12, 10, 30, 15, 15]
    for i, w in enumerate(col_widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ========================================
    # Sheet 4: Statistics (통계)
    # ========================================
    ws4 = wb.create_sheet("Statistics")

    stats = [
        ("PIPE BOM STATISTICS", ""),
        ("", ""),
        ("Total Pages", len(pages_data)),
        ("Total Pipe Pieces", sum(len(p["pipe_pieces"]) for p in pages_data)),
        ("Unique Base Pieces", len(piece_summary)),
        ("", ""),
        ("WELD SUMMARY", ""),
        ("Total Shop Welds", grand_shop),
        ("Total Field Welds (FFW)", grand_field),
        ("Total Welds", grand_total),
        ("", ""),
        ("PIPE LENGTH SUMMARY", ""),
        ("Total Measured Length (mm)", grand_length),
        ("Total Measured Length (m)", round(grand_length / 1000, 2)),
        ("", ""),
        ("LOOSE PARTS", ""),
        ("Pages with Loose Parts", sum(1 for p in pages_data if p.get("has_loose"))),
        ("", ""),
        ("REVISION NOTES", ""),
    ]

    # Add revision notes
    for pd in pages_data:
        for note in pd.get("revision_notes", []):
            stats.append((f"Page {pd['page']}", note))

    for r, (label, value) in enumerate(stats, 1):
        ws4.cell(row=r, column=1, value=label)
        ws4.cell(row=r, column=2, value=value)
        if label and not value and label == label.upper():
            ws4.cell(row=r, column=1).font = Font(name="Arial", size=11, bold=True)
        else:
            ws4.cell(row=r, column=1).font = Font(name="Arial", size=10)
            ws4.cell(row=r, column=2).font = Font(name="Arial", size=10)

    ws4.column_dimensions["A"].width = 35
    ws4.column_dimensions["B"].width = 50

    # Save
    wb.save(str(OUTPUT_PATH))
    print(f"\nExcel saved: {OUTPUT_PATH}")
    print(f"  Sheet 1: Pipe Piece Summary ({len(pages_data)} pages)")
    print(f"  Sheet 2: Weld Item Detail ({grand_total} items)")
    print(f"  Sheet 3: Weld Quantity Summary ({len(piece_summary)} base pieces)")
    print(f"  Sheet 4: Statistics")


if __name__ == "__main__":
    main()
