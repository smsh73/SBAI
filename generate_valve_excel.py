#!/usr/bin/env python3
"""VALVE-LIST Excel 생성 스크립트 - P&ID에서 추출한 밸브 데이터를 양식에 맞게 출력"""
import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from copy import copy
from pathlib import Path

SBAI_DIR = Path(__file__).parent
TEMPLATE_PATH = SBAI_DIR / "2. 260210-VALVE-LIST-양식-외부송부용.xlsx"
VALVE_DATA_PATH = SBAI_DIR / "valve_data.json"
OUTPUT_PATH = SBAI_DIR / "2. 260210-VALVE-LIST-OUTPUT.xlsx"

# Piping class → material mapping (from sample data in template)
PIPING_CLASS_MAP = {
    "CS3": {"piping_spec": "ACS10B3", "body": "ASTM A536", "trim": "B62",
            "flange": "150 Lbs RF WN\n(A105)", "pipe_mat": "ASTM A53 Gr.B _ POLYETHYLENE LINING INSIDE"},
    "CS2": {"piping_spec": "ACS10B2", "body": "ASTM A536", "trim": "B62",
            "flange": "150 Lbs RF WN\n(A105)", "pipe_mat": "ASTM A53 Gr.B _ POLYETHYLENE LINING INSIDE"},
}

# Special cases for SSW (Spray Sea Water) - from template sample row 8
SSW_SPEC = {
    "piping_spec": "BCS21B3", "body": "ASTM A536", "trim": "B62",
    "flange": "150 Lbs RF WN\n(A105)", "pipe_mat": "ASTM A53 Gr.B _ POLYETHYLENE LINING INSIDE"
}

# Schedule mapping from line spec
SCH_MAP = {
    "STD": "STD", "40": "40", "80": "80", "XS": "XS",
    "160": "160", "10": "10", "20": "20",
}

# Design pressure/temp by fluid type (from template samples)
DESIGN_CONDITIONS = {
    "SW": {"press": 6.5, "temp": 60},
    "SSW": {"press": 10, "temp": 60},
    "CFW": {"press": 6.5, "temp": 90},
    "FW": {"press": 6.5, "temp": 60},
}


def get_piping_spec(valve):
    """밸브의 piping spec 코드 결정"""
    fluid = valve.get("fluid", "")
    piping_class = valve.get("piping_class", "CS3")
    size = valve.get("size", "")

    # SSW 계열
    if fluid == "SW" and valve["tag"].startswith("SSW"):
        return SSW_SPEC["piping_spec"]

    cls_info = PIPING_CLASS_MAP.get(piping_class, PIPING_CLASS_MAP["CS3"])
    return cls_info["piping_spec"]


def get_material_info(valve):
    """밸브 재질 정보 결정"""
    piping_class = valve.get("piping_class", "CS3")
    fluid = valve.get("fluid", "")

    if fluid == "SW" and valve["tag"].startswith("SSW"):
        return SSW_SPEC
    return PIPING_CLASS_MAP.get(piping_class, PIPING_CLASS_MAP["CS3"])


def get_design_conditions(valve):
    """설계 압력/온도 결정"""
    fluid = valve.get("fluid", "")
    tag = valve.get("tag", "")

    if tag.startswith("SSW"):
        return DESIGN_CONDITIONS.get("SSW", {"press": 10, "temp": 60})

    return DESIGN_CONDITIONS.get(fluid, {"press": 6.5, "temp": 60})


def get_pipe_material(valve):
    """파이프 재질"""
    sch = valve.get("schedule", "STD")
    size = int(valve.get("size", "10")) if valve.get("size", "").isdigit() else 10

    mat_info = get_material_info(valve)
    base_mat = mat_info["pipe_mat"]

    # 2" 이하의 경우 GALVANIZING 적용 (template sample에서 CSW9109 참고)
    if size <= 2 and "GALVANIZING" not in base_mat:
        base_mat = "ASTM A53 Gr.B _ GALVANIZING"

    return base_mat


def main():
    # Load valve data
    with open(VALVE_DATA_PATH) as f:
        valves = json.load(f)

    print(f"Loaded {len(valves)} valves from JSON")

    # Load template
    wb = openpyxl.load_workbook(str(TEMPLATE_PATH))
    ws = wb["Manual"]

    # Get reference cell styles from row 8
    ref_styles = {}
    for cell in ws[8]:
        ref_styles[cell.column] = {
            "font": copy(cell.font),
            "alignment": copy(cell.alignment),
            "border": copy(cell.border),
            "fill": copy(cell.fill),
            "number_format": cell.number_format,
        }

    # Categorize valves
    manual_valves = [v for v in valves if v["valve_type"] not in ("CONTROL",)]
    control_valves = [v for v in valves if v["valve_type"] == "CONTROL"]

    # Sort: manual by tag, control by tag
    manual_valves.sort(key=lambda v: v["tag"])
    control_valves.sort(key=lambda v: v["tag"])

    # Clear existing data rows (keep header rows 1-7)
    for row_num in range(8, ws.max_row + 1):
        for col in range(1, 27):  # A to Z
            ws.cell(row=row_num, column=col).value = None

    # Write manual valves section header
    row = 7
    ws.cell(row=row, column=1, value="1. Manual Valve")

    # Write manual valves
    for i, valve in enumerate(manual_valves):
        row = 8 + i
        mat_info = get_material_info(valve)
        design = get_design_conditions(valve)
        sch = valve.get("schedule", "STD")

        ws.cell(row=row, column=1, value=i + 1)  # NO.
        ws.cell(row=row, column=2, value=valve["tag"])  # TAG
        ws.cell(row=row, column=3, value=get_piping_spec(valve))  # PIPING SPEC
        ws.cell(row=row, column=4, value=valve["location"])  # LOCATION
        ws.cell(row=row, column=5, value=valve["description"])  # DESCRIPTION
        ws.cell(row=row, column=6, value=valve["fluid"])  # FLUID
        ws.cell(row=row, column=7, value=design["press"])  # DESIGN PRESS
        ws.cell(row=row, column=8, value=design["temp"])  # DESIGN TEMP
        ws.cell(row=row, column=9, value=valve["valve_type"])  # VALVE TYPE
        ws.cell(row=row, column=10, value=valve.get("valve_subtype", "").replace(
            valve["valve_type"] + " ", "").strip() or valve["valve_type"])  # SUBTYPE
        ws.cell(row=row, column=11, value=mat_info["body"])  # MATERIAL BODY
        ws.cell(row=row, column=12, value=mat_info["trim"])  # MATERIAL TRIM
        ws.cell(row=row, column=13, value=int(valve["size"]) if valve["size"].isdigit() else valve["size"])  # SIZE
        ws.cell(row=row, column=14, value="ANSI")  # PRESSURE RATING CODE
        ws.cell(row=row, column=15, value=mat_info["flange"])  # FLANGE
        ws.cell(row=row, column=16, value="FLG")  # END CONN INLET
        ws.cell(row=row, column=17, value="FLG")  # END CONN OUTLET
        ws.cell(row=row, column=18, value=get_pipe_material(valve))  # PIPE MATERIAL
        ws.cell(row=row, column=19, value=sch)  # SCH IN
        ws.cell(row=row, column=20, value=sch)  # SCH OUT
        ws.cell(row=row, column=21, value="-")  # EXTENSION BONNET
        ws.cell(row=row, column=22, value=3)  # CLASS CERT
        ws.cell(row=row, column=23, value="-")  # LIMIT SWITCH OPEN
        ws.cell(row=row, column=24, value="-")  # LIMIT SWITCH CLOSE
        ws.cell(row=row, column=25, value="-")  # LOCK'G DEVICE
        ws.cell(row=row, column=26, value=f"Sheet {valve['sheet']}")  # REMARK

        # Apply styles from reference row
        for col in range(1, 27):
            cell = ws.cell(row=row, column=col)
            if col in ref_styles:
                cell.font = copy(ref_styles[col]["font"])
                cell.alignment = copy(ref_styles[col]["alignment"])
                cell.border = copy(ref_styles[col]["border"])

    # Write control valves section
    ctrl_start_row = 8 + len(manual_valves) + 1
    ws.cell(row=ctrl_start_row, column=1, value="2. Control Valve")
    ws.cell(row=ctrl_start_row, column=1).font = Font(name="Arial", size=10, bold=True)

    for i, valve in enumerate(control_valves):
        row = ctrl_start_row + 1 + i
        mat_info = get_material_info(valve)
        design = get_design_conditions(valve)
        sch = valve.get("schedule", "STD")

        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=valve["tag"])
        ws.cell(row=row, column=3, value=get_piping_spec(valve))
        ws.cell(row=row, column=4, value=valve["location"])
        ws.cell(row=row, column=5, value=valve["description"])
        ws.cell(row=row, column=6, value=valve["fluid"])
        ws.cell(row=row, column=7, value=design["press"])
        ws.cell(row=row, column=8, value=design["temp"])
        ws.cell(row=row, column=9, value="CONTROL")
        ws.cell(row=row, column=10, value=valve.get("valve_subtype", ""))
        ws.cell(row=row, column=11, value=mat_info["body"])
        ws.cell(row=row, column=12, value=mat_info["trim"])
        ws.cell(row=row, column=13, value=int(valve["size"]) if valve["size"].isdigit() else valve["size"])
        ws.cell(row=row, column=14, value="ANSI")
        ws.cell(row=row, column=15, value=mat_info["flange"])
        ws.cell(row=row, column=16, value="FLG")
        ws.cell(row=row, column=17, value="FLG")
        ws.cell(row=row, column=18, value=get_pipe_material(valve))
        ws.cell(row=row, column=19, value=sch)
        ws.cell(row=row, column=20, value=sch)
        ws.cell(row=row, column=21, value="-")
        ws.cell(row=row, column=22, value=3)
        ws.cell(row=row, column=23, value="-")
        ws.cell(row=row, column=24, value="-")
        ws.cell(row=row, column=25, value="-")
        ws.cell(row=row, column=26, value=f"Sheet {valve['sheet']}")

        for col in range(1, 27):
            cell = ws.cell(row=row, column=col)
            if col in ref_styles:
                cell.font = copy(ref_styles[col]["font"])
                cell.alignment = copy(ref_styles[col]["alignment"])
                cell.border = copy(ref_styles[col]["border"])

    # Summary row
    summary_row = ctrl_start_row + len(control_valves) + 2
    ws.cell(row=summary_row, column=1, value="TOTAL")
    ws.cell(row=summary_row, column=1).font = Font(name="Arial", size=10, bold=True)
    ws.cell(row=summary_row, column=2, value=f"Manual: {len(manual_valves)}, Control: {len(control_valves)}, Total: {len(valves)}")
    ws.cell(row=summary_row, column=2).font = Font(name="Arial", size=10, bold=True)

    # Save
    wb.save(str(OUTPUT_PATH))
    print(f"\nExcel saved: {OUTPUT_PATH}")
    print(f"  Manual valves: {len(manual_valves)}")
    print(f"  Control valves: {len(control_valves)}")
    print(f"  Total: {len(valves)}")

    # Print summary table
    print("\n=== VALVE TYPE SUMMARY ===")
    type_counts = {}
    for v in valves:
        vt = v["valve_type"]
        type_counts[vt] = type_counts.get(vt, 0) + 1
    for vt, cnt in sorted(type_counts.items(), key=lambda x: -x[1]):
        print(f"  {vt:20s}: {cnt}")

    print("\n=== SIZE SUMMARY ===")
    size_counts = {}
    for v in valves:
        s = v["size"]
        size_counts[s] = size_counts.get(s, 0) + 1
    for s, cnt in sorted(size_counts.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 0):
        print(f"  {s:>4s}\": {cnt}")


if __name__ == "__main__":
    main()
